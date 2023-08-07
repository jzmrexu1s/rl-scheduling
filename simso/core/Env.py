import math
import random
import numpy as np
import sys
sys.path.append("..")
from simso.core.Criticality import Criticality
import openpyxl
from simso.utils.norm import norm

profile = str(3)
workbook_title = 'Run 35_ LKATestBenchExample'

class Env(object):
    def __init__(self, model=None):
        self._model = model
        
        self.column_to_id = {
            "a_ego": 2, "status": 3, "v_ego": 5, "v_lead": 6, "a_lead": 7, "safe_distance": 8
        }
        self.column_inteval = {
            "ACC_status": 0.1, "a_ego": 0.1, "v_ego": 0.01, "v_lead": 0.01, "a_lead": 0.1, "safe_distance": 0.01,
            "LKA_status": 0.1, "steering_angle": 0.1, "curvature": 0.1, "headingAngle": 0.1, "lateralOffset": 0.1, "strength": 0.1,
            "departure_detected": 0.1, "lateral_deviation": 0.1
        }
        removed = {
            "status": 0.1, "driver_steer": 0.01
        }
        self.hi_task_name_to_column = {
            "T1": "ACC_status", "T2": "LKA_status"
        }
        
        self.sheets = {}
        for label in self.column_inteval.keys():
            workbook = openpyxl.load_workbook('./mpc/' + str(profile) + '/' + label + '.xlsx')
            sheet = workbook[workbook_title]
            self.sheets[label] = sheet
            
    def format_row(self, second, interval):
        row = 2
        if interval == 0.1:
            row += (int(second * 10) / 10) * 10
        if interval == 0.01:
            row += (int(second * 100) / 100) * 100
        return row
    
    def read_from_file_new(self, label, time):
        
        second = time / (1000000 * 1000)
        
        column = 2

        interval = self.column_inteval[label]
        row = self.format_row(second, interval)
        
        value = self.sheets[label].cell(row=int(row), column=column).value
        
        if value is None:
            prev = value
            p = second
            while self.format_row(p, interval) >= 2 and prev is None:
                p -= interval
                prev = self.sheets[label].cell(row=int(self.format_row(p, interval)), column=column).value
            next = value
            p = second
            while self.format_row(p, interval) < self.sheets[label].max_row and next is None:
                p += interval
                next = self.sheets[label].cell(row=int(self.format_row(p, interval)), column=column).value
            if prev is not None and next is not None:
                return (prev + next) / 2
            return prev
        
        return value
            
        
    def read_from_file(self, label, time):
        second = time / (1000000 * 1000)
        
        column = self.column_to_id[label]
        row = 2
        interval = self.column_inteval[label]
        
        if interval == 0.1:
            row += (int(second * 10) / 10) * 10
        if interval == 0.01:
            row += (int(second * 100) / 100) * 100
        return self.sheet.cell(row=int(row), column=column).value
        

    def observe_norm(self, time):
        # a_ego, v_ego, a_lead, v_lead, safe_distance
        a_ego = self.read_from_file_new("a_ego", time)
        v_ego = self.read_from_file_new("v_ego", time)
        a_lead = self.read_from_file_new("a_lead", time)
        v_lead = self.read_from_file_new("v_lead", time)
        safe_distance = self.read_from_file_new("safe_distance", time)
        # driver_steer = self.read_from_file_new("driver_steer", time)
        # curvature = self.read_from_file_new("curvature", time)
        headingAngle = self.read_from_file_new("headingAngle", time)
        lateralOffset = self.read_from_file_new("lateralOffset", time)
        strength = self.read_from_file_new("strength", time)
        departure_detected = self.read_from_file_new("departure_detected", time)
        lateral_deviation = self.read_from_file_new("lateral_deviation", time)
        
        # print(time, a_ego, v_ego, a_lead, v_lead, safe_distance)
        
        return np.array([
            norm(a_ego, -3, 2), norm(v_ego, 10, 40), norm(a_lead, -1, 1), norm(v_lead, 10, 40), norm(safe_distance, 30, 60),
            norm(headingAngle, -0.3, 0.3), norm(lateralOffset, -0.1, 3), norm(strength, 0.35, 0.65), norm(departure_detected, 0, 1), 
            norm(lateral_deviation, -0.5, 0.5)
        ])
    
    def now_acet(self, job, time):
        name = job.task.name
        # ms
        if job.task.criticality == Criticality.HI and name in ['T1', 'T2']:
            # 2ms一步
            return self.read_from_file_new(self.hi_task_name_to_column[name], time) * 3
        else:
            return math.ceil(100 * job.wcet * random.uniform(0.7, 1)) / 100
    
if __name__ == '__main__':
    env = Env()
    print(env.observe_norm(1668994370))